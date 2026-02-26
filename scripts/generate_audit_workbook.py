#!/usr/bin/env python3
"""
Generate a Data Quality Audit Workbook from grace_products_clean.json

This creates an Excel file with:
  - Tab 1: MASTER AUDIT — every SKU with flag columns
  - Tab 2: GLASS BOTTLES — filtered to Glass Bottle category only
  - Tab 3: THREAD SIZE REVIEW — grouped by family+capacity → verify thread
  - Tab 4: BOTTLE COLORS REVIEW — grouped by family+capacity → verify colors
  - Tab 5: CAP COLORS REVIEW — all cap colors for standardization
  - Tab 6: SUMMARY DASHBOARD — issue counts and quick stats
"""

import json
import os
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ─── CONFIG ───
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
DATA_FILE = os.path.join(PROJECT_DIR, "data", "grace_products_clean.json")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "docs", "PRODUCT_DATA_AUDIT.xlsx")

# ─── STYLES ───
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
FLAG_FILL_RED = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
FLAG_FILL_YELLOW = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
FLAG_FILL_GREEN = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
CORRECT_FONT = Font(name="Calibri", color="155724")
ERROR_FONT = Font(name="Calibri", color="721C24", bold=True)
WARN_FONT = Font(name="Calibri", color="856404")
SECTION_FILL = PatternFill(start_color="E8EDF3", end_color="E8EDF3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ─── KNOWN RULES ───
# Threading: mm-based names → standard equivalents
THREAD_STANDARDIZATION = {
    "13mm": "13-415",
    "16mm": "16-415",
    "16.5mm": "16-415",
    "18mm": "18-415",
}

# Suspect single-count colors (across the catalog)
SUSPECT_COLORS = {
    "Shiny", "Matte Shiny Silver", "Gold Silver", "Matte Red",
    "Copper Red", "Gold Ivory", "Gold Lavender Pink", "Copper Gold",
    "Black Copper", "Shiny Black Copper", "Shiny Copper Pink",
    "Shiny Black White", "Clear Gold", "Clear Silver",
}


def load_data():
    with open(DATA_FILE, "r") as f:
        return json.load(f)


def flag_product(p, all_products):
    """Generate flag columns for a single product."""
    flags = []
    flag_details = []
    severity = "✅"  # default

    family = p.get("family", "")
    cap_ml = p.get("capacityMl", 0)
    color = str(p.get("color", ""))
    thread = str(p.get("neckThreadSize", ""))
    applicator = p.get("applicator")
    cap_color = p.get("capColor")

    # ── Flag 1: Thread size standardization ──
    if thread in THREAD_STANDARDIZATION:
        flags.append("THREAD_NONSTANDARD")
        flag_details.append(
            f'Thread "{thread}" should be "{THREAD_STANDARDIZATION[thread]}"'
        )
        severity = "⚠️"

    # ── Flag 2: Rare bottle color ──
    color_count = sum(1 for pp in all_products if str(pp.get("color", "")) == color)
    if color and color_count < 5:
        flags.append("RARE_COLOR")
        flag_details.append(f'Color "{color}" only appears in {color_count} SKU(s)')
        severity = "⚠️"

    # ── Flag 3: Missing applicator ──
    if not applicator or applicator == "None":
        flags.append("MISSING_APPLICATOR")
        flag_details.append("No applicator specified")

    # ── Flag 4: Missing cap color ──
    if not cap_color or cap_color == "None":
        flags.append("MISSING_CAP_COLOR")
        flag_details.append("No cap color specified")

    # ── Flag 5: Missing thread size ──
    if not thread or thread == "None":
        flags.append("MISSING_THREAD")
        flag_details.append("No thread size specified")
        severity = "🔴"

    # ── Flag 6: Suspect color name ──
    if color in SUSPECT_COLORS:
        flags.append("SUSPECT_COLOR_NAME")
        flag_details.append(f'Color "{color}" looks like a data entry error')
        severity = "🔴"

    if cap_color and str(cap_color) in SUSPECT_COLORS:
        flags.append("SUSPECT_CAP_COLOR")
        flag_details.append(f'Cap color "{cap_color}" looks like a data entry error')
        severity = "🔴"

    # ── Flag 7: Multiple threads for this family+capacity ──
    group_threads = set()
    for pp in all_products:
        if pp.get("family") == family and pp.get("capacityMl") == cap_ml:
            t = str(pp.get("neckThreadSize", ""))
            if t and t != "None":
                group_threads.add(t)
    if len(group_threads) > 1 and family not in [
        "Component", "Cap", "Cap/Closure", "Dropper", "Sprayer",
        "Roll-On Cap", "Lotion Pump",
    ]:
        flags.append("MULTI_THREAD")
        flag_details.append(
            f"Family {family} {cap_ml}ml has multiple threads: {sorted(group_threads)}"
        )
        severity = "🔴"

    flag_count = len(flags)
    flag_str = "; ".join(flags) if flags else "CLEAN"
    detail_str = " | ".join(flag_details) if flag_details else "No issues"

    return flag_count, severity, flag_str, detail_str


def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            width = min(max(max(lengths), min_width), max_width)
            ws.column_dimensions[col_letter].width = width + 2


def create_master_audit(wb, products):
    ws = wb.active
    ws.title = "MASTER AUDIT"

    headers = [
        "Row #", "Grace SKU", "Website SKU", "Category", "Family", "Capacity (ml)",
        "Bottle Color", "Applicator", "Cap Color", "Thread Size",
        "Price 1pc", "Price 12pc", "Case Qty", "Stock Status",
        "# Flags", "Severity", "Flag Types", "Flag Details",
        "CORRECTED Color", "CORRECTED Thread", "CORRECTED Applicator",
        "CORRECTED Cap Color", "NOTES / ACTION"
    ]

    ws.append(headers)
    style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Pre-compute flags (this is O(n²) but fine for 2,354 rows)
    print("  Computing flags for all products...")
    flag_cache = {}
    for i, p in enumerate(products):
        flag_cache[i] = flag_product(p, products)

    print("  Writing rows...")
    for i, p in enumerate(products):
        flag_count, severity, flag_str, detail_str = flag_cache[i]

        row = [
            i + 1,
            p.get("graceSku", ""),
            p.get("websiteSku", ""),
            p.get("category", ""),
            p.get("family", ""),
            p.get("capacityMl", ""),
            str(p.get("color", "") or ""),
            str(p.get("applicator", "") or ""),
            str(p.get("capColor", "") or ""),
            str(p.get("neckThreadSize", "") or ""),
            p.get("webPrice1pc", ""),
            p.get("webPrice12pc", ""),
            p.get("caseQuantity", ""),
            p.get("stockStatus", ""),
            flag_count,
            severity,
            flag_str,
            detail_str,
            "",  # Corrected color
            "",  # Corrected thread
            "",  # Corrected applicator
            "",  # Corrected cap color
            "",  # Notes
        ]
        ws.append(row)

        # Color code the row based on severity
        row_num = i + 2
        if severity == "🔴":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = FLAG_FILL_RED
        elif severity == "⚠️":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = FLAG_FILL_YELLOW

        # Highlight the correction columns in light green so they stand out
        for col in range(19, 24):
            ws.cell(row=row_num, column=col).fill = FLAG_FILL_GREEN

    auto_width(ws)
    print(f"  → MASTER AUDIT: {len(products)} rows written")


def create_glass_bottles_tab(wb, products):
    ws = wb.create_sheet("GLASS BOTTLES")

    glass = [p for p in products if p.get("category") == "Glass Bottle"]

    headers = [
        "Grace SKU", "Family", "Capacity (ml)", "Bottle Color",
        "Applicator", "Cap Color", "Thread Size",
        "Price 1pc", "Stock Status",
        "Color OK?", "Thread OK?", "Applicator OK?", "Cap OK?", "NOTES"
    ]

    ws.append(headers)
    style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Sort by family → capacity → color → applicator
    glass.sort(key=lambda p: (
        p.get("family", ""),
        p.get("capacityMl", 0),
        str(p.get("color", "")),
        str(p.get("applicator", "")),
    ))

    for p in glass:
        row = [
            p.get("graceSku", ""),
            p.get("family", ""),
            p.get("capacityMl", ""),
            str(p.get("color", "") or ""),
            str(p.get("applicator", "") or ""),
            str(p.get("capColor", "") or ""),
            str(p.get("neckThreadSize", "") or ""),
            p.get("webPrice1pc", ""),
            p.get("stockStatus", ""),
            "",  # Color OK?
            "",  # Thread OK?
            "",  # Applicator OK?
            "",  # Cap OK?
            "",  # Notes
        ]
        ws.append(row)

    # Highlight validation columns
    for row_num in range(2, len(glass) + 2):
        for col in range(10, 15):
            ws.cell(row=row_num, column=col).fill = FLAG_FILL_GREEN

    auto_width(ws)
    print(f"  → GLASS BOTTLES: {len(glass)} rows written")


def create_thread_review(wb, products):
    ws = wb.create_sheet("THREAD SIZE REVIEW")

    headers = [
        "Family", "Capacity (ml)", "# SKUs",
        "Thread Sizes Found", "# Threads",
        "CORRECT Thread Size", "NOTES"
    ]

    ws.append(headers)
    style_header(ws, 1, len(headers))

    # Group by family + capacity
    groups = defaultdict(lambda: {"threads": set(), "count": 0})
    for p in products:
        key = (p.get("family", "?"), p.get("capacityMl", 0))
        t = str(p.get("neckThreadSize", "") or "")
        if t:
            groups[key]["threads"].add(t)
        groups[key]["count"] += 1

    for (fam, cap), data in sorted(groups.items()):
        threads = sorted(data["threads"])
        row = [
            fam,
            cap,
            data["count"],
            ", ".join(threads),
            len(threads),
            "",  # Correct thread
            "",  # Notes
        ]
        ws.append(row)

        row_num = ws.max_row
        if len(threads) > 1:
            ws.cell(row=row_num, column=4).fill = FLAG_FILL_RED
            ws.cell(row=row_num, column=5).fill = FLAG_FILL_RED
        ws.cell(row=row_num, column=6).fill = FLAG_FILL_GREEN
        ws.cell(row=row_num, column=7).fill = FLAG_FILL_GREEN

    auto_width(ws)
    print(f"  → THREAD SIZE REVIEW: {len(groups)} groups written")


def create_color_review(wb, products):
    ws = wb.create_sheet("BOTTLE COLORS REVIEW")

    headers = [
        "Family", "Capacity (ml)", "# SKUs",
        "Bottle Colors Found", "# Colors",
        "CORRECT Colors (comma-separated)", "COLORS TO REMOVE", "NOTES"
    ]

    ws.append(headers)
    style_header(ws, 1, len(headers))

    # Group by family + capacity
    groups = defaultdict(lambda: {"colors": Counter(), "count": 0})
    for p in products:
        key = (p.get("family", "?"), p.get("capacityMl", 0))
        c = str(p.get("color", "") or "?")
        groups[key]["colors"][c] += 1
        groups[key]["count"] += 1

    for (fam, cap), data in sorted(groups.items()):
        color_strs = [f"{c} ({ct})" for c, ct in data["colors"].most_common()]
        row = [
            fam,
            cap,
            data["count"],
            ", ".join(color_strs),
            len(data["colors"]),
            "",  # Correct colors
            "",  # Colors to remove
            "",  # Notes
        ]
        ws.append(row)

        row_num = ws.max_row
        # Flag groups with lots of rare colors
        rare = sum(1 for c, ct in data["colors"].items() if ct <= 2)
        if rare > 0:
            ws.cell(row=row_num, column=4).fill = FLAG_FILL_YELLOW

        ws.cell(row=row_num, column=6).fill = FLAG_FILL_GREEN
        ws.cell(row=row_num, column=7).fill = FLAG_FILL_GREEN
        ws.cell(row=row_num, column=8).fill = FLAG_FILL_GREEN

    auto_width(ws)
    print(f"  → BOTTLE COLORS REVIEW: {len(groups)} groups written")


def create_cap_color_review(wb, products):
    ws = wb.create_sheet("CAP COLORS REVIEW")

    headers = [
        "Cap Color (Current)", "# SKUs Using This",
        "Families Using This", "Category",
        "STANDARDIZED Name", "MERGE INTO", "NOTES"
    ]

    ws.append(headers)
    style_header(ws, 1, len(headers))

    # Collect cap color stats
    cap_stats = defaultdict(lambda: {"count": 0, "families": set()})
    for p in products:
        cc = str(p.get("capColor", "") or "")
        if cc:
            cap_stats[cc]["count"] += 1
            cap_stats[cc]["families"].add(p.get("family", "?"))

    for cc, data in sorted(cap_stats.items()):
        # Determine base color for grouping
        base = cc.replace("Matte ", "").replace("Shiny ", "").replace("Light ", "").strip()

        row = [
            cc,
            data["count"],
            ", ".join(sorted(data["families"])),
            base,
            "",  # Standardized name
            "",  # Merge into
            "",  # Notes
        ]
        ws.append(row)

        row_num = ws.max_row
        if data["count"] <= 2:
            ws.cell(row=row_num, column=1).fill = FLAG_FILL_RED
        elif cc in SUSPECT_COLORS:
            ws.cell(row=row_num, column=1).fill = FLAG_FILL_YELLOW

        ws.cell(row=row_num, column=5).fill = FLAG_FILL_GREEN
        ws.cell(row=row_num, column=6).fill = FLAG_FILL_GREEN
        ws.cell(row=row_num, column=7).fill = FLAG_FILL_GREEN

    auto_width(ws)
    print(f"  → CAP COLORS REVIEW: {len(cap_stats)} unique colors written")


def create_summary(wb, products):
    ws = wb.create_sheet("SUMMARY")

    # ── Title ──
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "📊 DATA QUALITY AUDIT SUMMARY"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1B2A4A")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1).value = f"Generated from grace_products_clean.json — {len(products)} SKUs"
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # ── Stats ──
    row = 4
    stats = [
        ("Total SKUs", len(products)),
        ("Glass Bottles", sum(1 for p in products if p.get("category") == "Glass Bottle")),
        ("Lotion Bottles", sum(1 for p in products if p.get("category") == "Lotion Bottle")),
        ("Components", sum(1 for p in products if p.get("category") == "Component")),
        ("Other Categories", sum(1 for p in products if p.get("category") not in ["Glass Bottle", "Lotion Bottle", "Component"])),
        ("", ""),
        ("Unique Families", len(set(p.get("family") for p in products))),
        ("Unique Bottle Colors", len(set(str(p.get("color", "")) for p in products if p.get("color")))),
        ("Unique Applicators", len(set(str(p.get("applicator", "")) for p in products if p.get("applicator")))),
        ("Unique Cap Colors", len(set(str(p.get("capColor", "")) for p in products if p.get("capColor")))),
        ("Unique Thread Sizes", len(set(str(p.get("neckThreadSize", "")) for p in products if p.get("neckThreadSize")))),
        ("", ""),
        ("⚠️ FLAGS FOUND", ""),
        ("SKUs with missing applicator", sum(1 for p in products if not p.get("applicator"))),
        ("SKUs with missing cap color", sum(1 for p in products if not p.get("capColor"))),
        ("SKUs with missing thread size", sum(1 for p in products if not p.get("neckThreadSize"))),
        ("SKUs with rare bottle color (<5 total)", "see BOTTLE COLORS tab"),
        ("Product groups with multiple threads", "see THREAD SIZE tab"),
        ("", ""),
        ("🔧 HOW TO USE THIS WORKBOOK", ""),
        ("1.", "Go to each review tab (THREAD SIZE, BOTTLE COLORS, CAP COLORS)"),
        ("2.", "Fill in the GREEN correction columns with the correct values"),
        ("3.", "For MASTER AUDIT, use the CORRECTED columns or add NOTES"),
        ("4.", "Filter by 'Severity' column to focus on 🔴 issues first"),
        ("5.", "When done, save and share — we'll use this to generate a clean data import"),
    ]

    for label, value in stats:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True if label.startswith(("⚠️", "🔧", "Total")) else False, size=11)
        ws.cell(row=row, column=2).value = value
        if label.startswith(("1.", "2.", "3.", "4.", "5.")):
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, color="1B2A4A")
        row += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 50
    print(f"  → SUMMARY: dashboard created")


def main():
    print("=" * 60)
    print("  GENERATING PRODUCT DATA AUDIT WORKBOOK")
    print("=" * 60)

    print(f"\n📂 Reading: {DATA_FILE}")
    products = load_data()
    print(f"   Loaded {len(products)} products\n")

    wb = Workbook()

    print("📝 Building tabs...")
    create_master_audit(wb, products)
    create_glass_bottles_tab(wb, products)
    create_thread_review(wb, products)
    create_color_review(wb, products)
    create_cap_color_review(wb, products)
    create_summary(wb, products)

    # Move summary to first position
    summary_ws = wb["SUMMARY"]
    wb.move_sheet(summary_ws, offset=-5)

    print(f"\n💾 Saving: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Done! Open the file in Excel:")
    print(f"   {OUTPUT_FILE}")
    print()


if __name__ == "__main__":
    main()
