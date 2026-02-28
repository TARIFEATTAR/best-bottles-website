#!/usr/bin/env python3
"""
Generate CONVEX_PRODUCT_ALIGNMENT_WORKBOOK.xlsx from the data alignment templates.

Sheets:
  - Products Template: All product columns with metadata rows + empty data rows
  - Product Groups Template: All product group columns with metadata rows
  - Column Reference: Full reference (Convex vs UI) for gap analysis
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(PROJECT_DIR, "docs", "data_alignment")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "docs", "data_alignment", "CONVEX_PRODUCT_ALIGNMENT_WORKBOOK.xlsx")

HEADER_FONT = Font(bold=True, size=11)
META_FILL = PatternFill(start_color="E8EDF3", end_color="E8EDF3", fill_type="solid")


def load_csv(path):
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.reader(f))


def write_template_sheet(wb, name, rows, freeze_at=5):
    ws = wb.create_sheet(name)
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = HEADER_FONT
            elif r <= 4:
                cell.fill = META_FILL
    ws.freeze_panes = ws.cell(row=freeze_at, column=1)
    return ws


def main():
    os.makedirs(DATA_DIR, exist_ok=True)

    products = load_csv(os.path.join(DATA_DIR, "CONVEX_PRODUCT_ALIGNMENT_TEMPLATE.csv"))
    groups = load_csv(os.path.join(DATA_DIR, "CONVEX_PRODUCT_GROUP_TEMPLATE.csv"))
    ref = load_csv(os.path.join(DATA_DIR, "CONVEX_COLUMN_REFERENCE.csv"))

    wb = Workbook()
    wb.remove(wb.active)

    # Products: 4 metadata rows + 10 empty data rows
    product_rows = products + [[""] * len(products[0]) for _ in range(10)]
    write_template_sheet(wb, "Products Template", product_rows)

    # Product Groups
    group_rows = groups + [[""] * len(groups[0]) for _ in range(10)]
    write_template_sheet(wb, "Product Groups Template", group_rows)

    # Column Reference
    ref_ws = wb.create_sheet("Column Reference")
    for r, row in enumerate(ref, 1):
        for c, val in enumerate(row, 1):
            ref_ws.cell(row=r, column=c, value=val)
            if r == 1:
                ref_ws.cell(row=r, column=c).font = HEADER_FONT
    ref_ws.freeze_panes = ref_ws.cell(row=2, column=1)

    # Auto-fit columns (approximate)
    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(30, max(12, len(str(ws.cell(1, col).value or "")) + 2))

    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
